
from openpyxl import workbook
import requests
from bs4 import BeautifulSoup
import openpyxl as ox
from openpyxl.styles import PatternFill, Font, Color
from links import links

section_title_bgstyle = PatternFill(patternType='solid', fgColor='81F608')


head = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.106 Safari/537.36'}


xl = ox.load_workbook('dot.xlsx')
courses = xl['extra']

for column_count, link in enumerate(links):
    print("link", column_count+1, "Started")
    row_count = 3
    req_text = requests.get(link, headers=head).text
    soup = BeautifulSoup(req_text, 'lxml')
    lecture_sidebar = soup.find('div', class_="row lecture-sidebar")
    course_section = lecture_sidebar.find_all('div', class_="course-section")
    for section in course_section:
        section_title = section.find(
            'div', class_='section-title').text.strip()
        courses.cell(row_count, column_count+1).value = section_title
        courses.cell(row_count, column_count+1).fill = section_title_bgstyle
        print("Section heading success")
        row_count += 1
        video_link_list = section.find_all('li', class_='section-item')
        for vdolink in video_link_list:
            lecture_title = ' '.join(vdolink.find(
                'span', class_='lecture-name').text.strip().replace('\n', '').split())
            courses.cell(row_count, column_count+1).value = lecture_title
            path_link = vdolink.find('a', class_='item')['href']
            sub_req = requests.get(
                "https://codewithmosh.com" + path_link, headers=head).text
            sub_soup = BeautifulSoup(sub_req, 'lxml')
            download_link = sub_soup.find('a', class_="download")
            if download_link is not None:
                download_link = download_link['href'].strip()
                courses.cell(row_count, column_count +
                             1).hyperlink = download_link
                row_count += 1
            else:
                row_count += 1
        print("Section Completed")
xl.save('dot.xlsx')
