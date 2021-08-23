from typing import cast
from str_links import kkk
import openpyxl as ox
from openpyxl.styles import Font, Color

link_split = kkk.split('##l')
data_obj = {}
col_num = int(input("Colum Number  >>>>   "))
sub = input("having Sub Parts (y/n)  >>>>   ")
if sub == 'y':
    sub_con = 0
elif sub == 'n':
    sub_con = 1

xl_file_path = 'dot.xlsx'
selected_sheet = 'extra'
xl = ox.load_workbook(xl_file_path)
sheet = xl[selected_sheet]
for svid in link_split:
    spl = svid.split("##D")
    data_obj[spl[0].strip()] = spl[1]
alpha = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
         'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
cell_max_rows = len(
    [cell for cell in sheet[alpha[col_num - 1]] if cell.value is not None]) + sub_con
for row in range(3, cell_max_rows + 1):
    if sheet.cell(row, col_num).fill.fgColor.rgb == 'FF81F608':
        continue
    cell = sheet.cell(row, col_num)
    try:
        checklink = data_obj[cell.value.strip()]
    except KeyError:
        continue
    if cell.hyperlink == None:
        attach = data_obj[cell.value.strip()]
        cell.hyperlink = attach
xl.save('dot.xlsx')
