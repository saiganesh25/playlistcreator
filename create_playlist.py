from pathlib import Path
import openpyxl
from string import Template

i = 1

while i == 1:
    # xl sheet info
    xl_file_path = 'dot.xlsx'
    selected_sheet = 'back-end'
    xl = openpyxl.load_workbook(xl_file_path)
    sheet = xl[selected_sheet]

    def course(course_name, column_number, column_alpha, sub='n', sub_path=''):

        # -------------->>>>>>export Details here
        course = course_name
        extension = '.xspf'

        column_count = column_number
        course_sub_parts = 0 if sub == 'y' else 1
        column_name = column_alpha
        section_cell_color = 'FF81F608'
        disable_cell_color = 'FFCCCCCC'
        course_dir = f"courses/{course}"

        if not Path(course_dir).exists():
            Path(course_dir).mkdir()

        if sub == 'y':
            course_dir = course_dir + f"/{sub_path}"
            if not Path(course_dir).exists():
                Path(course_dir).mkdir()

        # -------------Templates-----------------------------

        single_vid_temp = Template(
            Path('temp/singlevideotemplate.xml').read_text())
        multi_vid_temp = Template(Path('temp/multiple_main.xml').read_text())
        multi_s1 = Template(Path('temp/multiple_s1.xml').read_text())
        multi_s2 = Template(Path('temp/multiple_s2.xml').read_text())

        cell_max_rows = len(
            [cell for cell in sheet[column_name] if cell.value is not None]) + course_sub_parts

        section_count = 0
        video_count = 0
        section_track_list = ''
        section_id_list = ''
        course_track_list = ''
        course_id_list = ''

        for row in range(3, cell_max_rows + 1):
            color_sec = sheet.cell(row, column_count).fill.fgColor.rgb
            if color_sec == section_cell_color:
                section_count += 1
                section_name = sheet.cell(
                    row, column_count).value.replace(':', ' -')
                section_folder_name = str(
                    section_count) + ' ' + section_name
                section_link = course_dir + f"/{section_folder_name}"
                Path(section_link).mkdir()
                section_videos_subfolder = section_link + '/section videos'
                Path(section_videos_subfolder).mkdir()
                continue
            if sheet.cell(row, column_count).fill.bgColor.rgb == disable_cell_color:
                continue
            vid_title = sheet.cell(row, column_count).value.strip()
            check_cout = vid_title.split()
            try:
                isinstance(int(check_cout[0].replace('-', '')), int)
            except ValueError:
                vid_title = str(video_count + 1) + '-' + ' ' + vid_title
            title_list = vid_title.split()
            if ':' in title_list[len(title_list)-1]:
                del title_list[len(title_list)-1]
            title = ' '.join(title_list).replace(':', '')
            title = title.replace('?', '')
            single_vid_file_name = title + extension
            link = sheet.cell(row, column_count).hyperlink
            if link is not None:
                target = link.target
            else:
                continue
                # if sub == 'n':
                #     target = '../../../asset/atm.jpg'
                # else:
                #     target = '../../../../asset/atm.jpg'

            template_code = single_vid_temp.substitute(
                {'title': title, 'link': target})
            vlc_file = Path(
                f"{section_link}/section videos/{single_vid_file_name}")
            vlc_file.write_text(template_code)

            section_track_list += multi_s1.substitute(
                {'link': (f'section videos/{single_vid_file_name}'), 'id':  str(video_count)})
            section_id_list += multi_s2.substitute(
                {'id': f'"{str(video_count)}"'})
            video_count += 1

            if (sheet.cell(row + 1, column_count).fill.bgColor.rgb == section_cell_color) or (row == cell_max_rows):
                if section_count == 0:
                    pass
                else:
                    section_vid_temp_code = multi_vid_temp.substitute(
                        {'tracks': section_track_list, 'idm': section_id_list})
                    section_vid_file_name = section_name + ' ' + \
                        ' -- Full Section Lecture' + extension
                    vlc_section_file = Path(
                        f"{section_link}/{section_vid_file_name}")
                    vlc_section_file.write_text(section_vid_temp_code)
                    section_track_list = ''
                    section_id_list = ''
                    video_count = 0
                # ------------------------------------------------
                    course_track_list += multi_s1.substitute(
                        {'link': (section_folder_name + '/' + section_vid_file_name), 'id': section_count - 1})

                    course_id_list += multi_s2.substitute(
                        {'id': f'"{str(section_count - 1)}"'})

        full_course_template_code = multi_vid_temp.substitute(
            {'tracks': course_track_list, 'idm': course_id_list})
        full_course_file_name = course + ' -- Full Lecture' + extension
        full_course_vlc = Path(
            course_dir + '/' + full_course_file_name).write_text(full_course_template_code)
        print("\nProcess Completed Successfully !!!\n")
        i = 0


# ---------------------------------------------------------------------------------------------------------------

    cstart = int(input("\nColumn Start >>>>>  "))
    cname = sheet.cell(1, cstart).value

    if cname is not None:
        cname = cname.strip()
        print(f"\nCreating playlist for {cname}, Do you want to proceed?\n")
        if input("Your Answer --> (y/n) >>>>> ") == 'y':
            if not Path(f'courses/{cname}').exists():
                spart = 'y' if sheet.cell(1, cstart + 1).value is None else 'n'
                alpha = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
                         'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

                dynamic_sub_count = 1

                if spart == 'y':
                    while True:
                        if sheet.cell(1, cstart + dynamic_sub_count).value is not None:
                            break
                        else:
                            if dynamic_sub_count > 10:
                                dynamic_sub_count = int(input(
                                    "\nSub parts reached the limit, please provide sub Count\n>>>>>      "))
                                break
                            dynamic_sub_count += 1

                if spart == 'n':
                    calpha = alpha[cstart-1]
                    course(course_name=cname, column_number=cstart,
                           column_alpha=calpha)
                elif spart == 'y':
                    total_subs = dynamic_sub_count
                    spart_name = []
                    for sub_row in range(total_subs):
                        spart_name.append(sheet.cell(
                            row=2, column=cstart+sub_row).value.strip())
                    for sub_num in spart_name:
                        cstart += sub_num
                        calpha = alpha[cstart-1]
                        spath = spart_name[sub_num]
                        course(course_name=cname, column_number=cstart,
                               column_alpha=calpha, sub=spart, sub_path=spath)
            else:
                print(
                    '\nthis Playlist already Created, To create New one Delete the existing Playlist!\n')
                i = 0
        else:
            print("\nProcess CANCELED Successfully....!\n")
            i = 0
    else:
        print("\nProcess CANCELED Due To selecting Wrong Column....!\n")
        i = int(input("Press 1 to proceed / 0 to Exit  >>>>>    "))
